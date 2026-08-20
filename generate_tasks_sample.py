import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_tasks_files():
    # 1. Generate CSV
    csv_content = """Client PAN,Task Title,Category,Priority,Due Date,Description
AAACZ1234D,GSTR-3B Monthly Return Filing & Tax Payment,GST,HIGH,2026-08-20,Reconciliation of GSTR-2B ITC and tax payment
AAALB5678E,GSTR-1 Monthly Outward Supplies Filing,GST,HIGH,2026-08-11,Monthly B2B invoices and B2CS sales summary upload
ABCPJ9876M,Annual ITR-2 Return Filing AY 2026-27,ITR,MEDIUM,2026-07-31,Salaried and capital gains AIS/TIS reconciliation
AABFM1122K,GSTR-3B Monthly Return Filing,GST,HIGH,2026-08-20,Monthly turnover tax computation and GSTR-3B return
AAACA4321C,Tax Audit Report Form 3CA/3CD Filing,AUDIT,URGENT,2026-09-30,Statutory tax audit under section 44AB
AABFS7788P,Quarterly TDS 26Q Return Q1,COMPLIANCE,HIGH,2026-07-31,Non-salary contractor and professional fee TDS
AHIPS3456L,ITR-1 Sahaj Return AY 2026-27,ITR,LOW,2026-07-31,Form 16 salary income tax return
AAALS9988Q,Quarterly Advance Tax Installment Q2,COMPLIANCE,HIGH,2026-09-15,Advance tax computation and challan generation
AAACM6655B,Annual ITR-6 Corporate Tax Return AY 2026-27,ITR,URGENT,2026-10-31,Corporate income tax return with MAT
AAATR1199F,Form 10B Audit Report & ITR-7,AUDIT,HIGH,2026-09-30,Educational trust statutory audit and exemption filing
AAAFS5544H,HUF Annual Tax Return Filing,ITR,MEDIUM,2026-07-31,Karta income computation and return submission
"""

    with open("d:/Projects/Taxoryn/sample_tasks_bulk_upload.csv", "w", encoding="utf-8") as f:
        f.write(csv_content)

    with open("d:/Projects/Taxoryn/frontend/public/sample_tasks_bulk_upload.csv", "w", encoding="utf-8") as f:
        f.write(csv_content)

    # 2. Generate Styled Excel Workbook (.xlsx)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Practice Tasks Bulk Import"

    headers = ["Client PAN", "Task Title", "Category", "Priority", "Due Date", "Description"]
    ws.append(headers)

    tasks_data = [
        ["AAACZ1234D", "GSTR-3B Monthly Return Filing & Tax Payment", "GST", "HIGH", "2026-08-20", "Reconciliation of GSTR-2B ITC and tax payment"],
        ["AAALB5678E", "GSTR-1 Monthly Outward Supplies Filing", "GST", "HIGH", "2026-08-11", "Monthly B2B invoices and B2CS sales summary upload"],
        ["ABCPJ9876M", "Annual ITR-2 Return Filing AY 2026-27", "ITR", "MEDIUM", "2026-07-31", "Salaried and capital gains AIS/TIS reconciliation"],
        ["AABFM1122K", "GSTR-3B Monthly Return Filing", "GST", "HIGH", "2026-08-20", "Monthly turnover tax computation and GSTR-3B return"],
        ["AAACA4321C", "Tax Audit Report Form 3CA/3CD Filing", "AUDIT", "URGENT", "2026-09-30", "Statutory tax audit under section 44AB"],
        ["AABFS7788P", "Quarterly TDS 26Q Return Q1", "COMPLIANCE", "HIGH", "2026-07-31", "Non-salary contractor and professional fee TDS"],
        ["AHIPS3456L", "ITR-1 Sahaj Return AY 2026-27", "ITR", "LOW", "2026-07-31", "Form 16 salary income tax return"],
        ["AAALS9988Q", "Quarterly Advance Tax Installment Q2", "COMPLIANCE", "HIGH", "2026-09-15", "Advance tax computation and challan generation"],
        ["AAACM6655B", "Annual ITR-6 Corporate Tax Return AY 2026-27", "ITR", "URGENT", "2026-10-31", "Corporate income tax return with MAT"],
        ["AAATR1199F", "Form 10B Audit Report & ITR-7", "AUDIT", "HIGH", "2026-09-30", "Educational trust statutory audit and exemption filing"],
        ["AAAFS5544H", "HUF Annual Tax Return Filing", "ITR", "MEDIUM", "2026-07-31", "Karta income computation and return submission"],
    ]

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    for row_idx, row_values in enumerate(tasks_data, start=2):
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if col_idx in [1, 3, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    wb.save("d:/Projects/Taxoryn/Taxoryn_Sample_Tasks_Bulk_Upload.xlsx")
    wb.save("d:/Projects/Taxoryn/frontend/public/Taxoryn_Sample_Tasks_Bulk_Upload.xlsx")
    print("Tasks sample Excel and CSV created successfully!")

if __name__ == "__main__":
    create_sample_tasks_files()
