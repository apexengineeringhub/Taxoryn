import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Practice Customers Migration"

    # Header Definition
    headers = [
        "Display Name",
        "Legal Name",
        "PAN",
        "GSTIN",
        "Client Type",
        "Email",
        "Phone",
        "City",
        "State",
        "Pincode"
    ]

    # Sample Data Rows
    data = [
        [
            "Zenith Infotech Solutions Pvt Ltd",
            "Zenith Infotech Solutions Private Limited",
            "AAACZ1234D",
            "27AAACZ1234D1Z8",
            "PRIVATE_LIMITED",
            "finance@zenithinfo.com",
            "9820011223",
            "Mumbai",
            "Maharashtra",
            "400021"
        ],
        [
            "Bluecrest Logistics LLP",
            "Bluecrest Logistics LLP",
            "AAALB5678E",
            "27AAALB5678E1Z4",
            "LLP",
            "accounts@bluecrestlog.com",
            "9820022334",
            "Pune",
            "Maharashtra",
            "411001"
        ],
        [
            "Anand Ramesh Joshi",
            "",
            "ABCPJ9876M",
            "",
            "INDIVIDUAL",
            "anand.joshi@gmail.com",
            "9820033445",
            "Nagpur",
            "Maharashtra",
            "440001"
        ],
        [
            "Mundeshwari Trading Co",
            "Mundeshwari Trading Proprietorship",
            "AABFM1122K",
            "27AABFM1122K1Z3",
            "PROPRIETORSHIP",
            "contact@mundeshwari.in",
            "9820044556",
            "Mumbai",
            "Maharashtra",
            "400001"
        ],
        [
            "Apex Agro Food Products Ltd",
            "Apex Agro Food Products Public Limited",
            "AAACA4321C",
            "27AAACA4321C1Z6",
            "PUBLIC_LIMITED",
            "tax@apexagro.co.in",
            "9820055667",
            "Nashik",
            "Maharashtra",
            "422001"
        ],
        [
            "Shree Ganesh Enterprises",
            "Shree Ganesh Enterprises Partnership",
            "AABFS7788P",
            "27AABFS7788P1Z9",
            "PARTNERSHIP",
            "tax@ganeshent.com",
            "9820066778",
            "Thane",
            "Maharashtra",
            "400601"
        ],
        [
            "Kavita Suresh Sharma",
            "",
            "AHIPS3456L",
            "",
            "INDIVIDUAL",
            "kavita.sharma@outlook.com",
            "9820077889",
            "Navi Mumbai",
            "Maharashtra",
            "400703"
        ],
        [
            "Sunrise Healthcare & Diagnostic Center",
            "Sunrise Healthcare LLP",
            "AAALS9988Q",
            "27AAALS9988Q1Z1",
            "LLP",
            "billing@sunrisehealth.org",
            "9820088990",
            "Aurangabad",
            "Maharashtra",
            "431001"
        ],
        [
            "Metro Infrastructure Developers Pvt Ltd",
            "Metro Infrastructure Developers Private Limited",
            "AAACM6655B",
            "27AAACM6655B1Z2",
            "PRIVATE_LIMITED",
            "accounts@metroinfra.in",
            "9820099001",
            "Mumbai",
            "Maharashtra",
            "400051"
        ],
        [
            "Royal Heritage Educational Trust",
            "Royal Heritage Educational Trust",
            "AAATR1199F",
            "",
            "TRUST",
            "trustee@royalheritage.edu.in",
            "9820100112",
            "Kolhapur",
            "Maharashtra",
            "416001"
        ],
        [
            "Singhania HUF",
            "Singhania Hindu Undivided Family",
            "AAAFS5544H",
            "",
            "HUF",
            "singhania.huf@gmail.com",
            "9820111223",
            "Mumbai",
            "Maharashtra",
            "400004"
        ]
    ]

    # Write Headers
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # Write Data Rows
    for row_idx, row_values in enumerate(data, start=2):
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if col_idx in [3, 4, 10]:  # PAN, GSTIN, Pincode
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    # Save to Root and Public
    wb.save("d:/Projects/Taxoryn/Taxoryn_Sample_Customers_Migration.xlsx")
    wb.save("d:/Projects/Taxoryn/frontend/public/Taxoryn_Sample_Customers_Migration.xlsx")
    print("Excel files created successfully!")

if __name__ == "__main__":
    create_sample_excel()
