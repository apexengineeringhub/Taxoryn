import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_staff_files():
    # 1. CSV File Content
    csv_content = """Employee Code,First Name,Last Name,Email,Phone,Department,Designation,Status
EMP-101,Rohan,Deshmukh,rohan.deshmukh@maamundeshwari.com,+919876543210,Taxation,Senior CA Partner / Practitioner,ACTIVE
EMP-102,Priya,Sharma,priya.sharma@maamundeshwari.com,+919812345678,GST & Indirect Tax,GST Filing Specialist,ACTIVE
EMP-103,Amit,Verma,amit.verma@maamundeshwari.com,+919823456789,Audit & Assurance,Audit Manager,ACTIVE
EMP-104,Sneha,Gupta,sneha.gupta@maamundeshwari.com,+919834567890,Direct Tax,Senior Tax Advocate,ACTIVE
EMP-105,Vikas,Patel,vikas.patel@maamundeshwari.com,+919845678901,Corporate Compliance,Company Secretary / Legal Associate,ACTIVE
EMP-106,Ananya,Sen,ananya.sen@maamundeshwari.com,+919856789012,Accounting & Bookkeeping,Senior Accountant,ACTIVE
EMP-107,Rahul,Mishra,rahul.mishra@maamundeshwari.com,+919867890123,Audit & Assurance,Article Assistant / Trainee,ACTIVE
EMP-108,Pooja,Joshi,pooja.joshi@maamundeshwari.com,+919878901234,Direct Tax,Article Assistant / Trainee,ACTIVE
"""

    with open("d:/Projects/Taxoryn/sample_staff_bulk_upload.csv", "w", encoding="utf-8") as f:
        f.write(csv_content)

    with open("d:/Projects/Taxoryn/frontend/public/sample_staff_bulk_upload.csv", "w", encoding="utf-8") as f:
        f.write(csv_content)

    # 2. Styled Excel Workbook (.xlsx)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Practice Staff Bulk Onboarding"

    headers = ["Employee Code", "First Name", "Last Name", "Email", "Phone", "Department", "Designation", "Status"]
    ws.append(headers)

    staff_data = [
        ["EMP-101", "Rohan", "Deshmukh", "rohan.deshmukh@maamundeshwari.com", "+919876543210", "Taxation", "Senior CA Partner / Practitioner", "ACTIVE"],
        ["EMP-102", "Priya", "Sharma", "priya.sharma@maamundeshwari.com", "+919812345678", "GST & Indirect Tax", "GST Filing Specialist", "ACTIVE"],
        ["EMP-103", "Amit", "Verma", "amit.verma@maamundeshwari.com", "+919823456789", "Audit & Assurance", "Audit Manager", "ACTIVE"],
        ["EMP-104", "Sneha", "Gupta", "sneha.gupta@maamundeshwari.com", "+919834567890", "Direct Tax", "Senior Tax Advocate", "ACTIVE"],
        ["EMP-105", "Vikas", "Patel", "vikas.patel@maamundeshwari.com", "+919845678901", "Corporate Compliance", "Company Secretary / Legal Associate", "ACTIVE"],
        ["EMP-106", "Ananya", "Sen", "ananya.sen@maamundeshwari.com", "+919856789012", "Accounting & Bookkeeping", "Senior Accountant", "ACTIVE"],
        ["EMP-107", "Rahul", "Mishra", "rahul.mishra@maamundeshwari.com", "+919867890123", "Audit & Assurance", "Article Assistant / Trainee", "ACTIVE"],
        ["EMP-108", "Pooja", "Joshi", "pooja.joshi@maamundeshwari.com", "+919878901234", "Direct Tax", "Article Assistant / Trainee", "ACTIVE"],
    ]

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    for row_idx, row_values in enumerate(staff_data, start=2):
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if col_idx in [1, 5, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    wb.save("d:/Projects/Taxoryn/Taxoryn_Sample_Staff_Bulk_Upload.xlsx")
    wb.save("d:/Projects/Taxoryn/frontend/public/Taxoryn_Sample_Staff_Bulk_Upload.xlsx")
    print("Staff sample Excel and CSV created successfully!")

if __name__ == "__main__":
    create_sample_staff_files()
