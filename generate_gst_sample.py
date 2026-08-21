import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_gst_files():
    # 1. GST Profiles CSV
    profiles_csv = """Client PAN,GSTIN,Legal Name,Trade Name,GST Scheme,Filing Frequency,Registration Date,State Code
AAACZ1234D,27AAACZ1234D1Z8,Zenith Infotech Solutions Private Limited,Zenith Infotech,REGULAR,MONTHLY,2018-04-01,27
AAALB5678E,29AAALB5678E1ZB,Bluecrest Logistics LLP,Bluecrest Express,REGULAR,MONTHLY,2019-07-15,29
ABCPJ9876M,27ABCPJ9876M1ZX,Anand Ramesh Joshi,Joshi Tax Consultancy,REGULAR,QUARTERLY,2021-11-01,27
AABFM1122K,27AABFM1122K1Z4,Mundeshwari Trading Co,Mundeshwari Traders,COMPOSITION,QUARTERLY,2017-07-01,27
AAACA4321C,24AAACA4321C1Z2,Apex Agro Food Products Limited,Apex Foods,REGULAR,MONTHLY,2018-09-20,24
AABFS7788P,07AABFS7788P1Z6,Shree Ganesh Enterprises,Ganesh Retailers,REGULAR,MONTHLY,2020-01-10,07
AAALS9988Q,27AAALS9988Q1Z9,Sunrise Healthcare LLP,Sunrise Clinic & Diagnostics,REGULAR,MONTHLY,2021-03-05,27
AAACM6655B,33AAACM6655B1Z1,Metro Infrastructure Developers Private Limited,Metro Infra,REGULAR,MONTHLY,2017-08-12,33
"""

    with open("d:/Projects/Taxoryn/sample_gst_profiles_migration.csv", "w", encoding="utf-8") as f:
        f.write(profiles_csv)
    with open("d:/Projects/Taxoryn/frontend/public/sample_gst_profiles_migration.csv", "w", encoding="utf-8") as f:
        f.write(profiles_csv)

    # 2. GST Filings CSV
    filings_csv = """GSTIN,Return Type,Return Period,Financial Year,Due Date,Filing Status,Taxable Value,Tax Liability,ITC Claimed,ARN Number
27AAACZ1234D1Z8,GSTR3B,2026-07,2026-27,2026-08-20,FILED,1850000,333000,240000,AA2707261234567
27AAACZ1234D1Z8,GSTR1,2026-07,2026-27,2026-08-11,FILED,1850000,333000,0,AA2707261234568
29AAALB5678E1ZB,GSTR3B,2026-07,2026-27,2026-08-20,FILED,920000,165600,110000,AA2907269876543
29AAALB5678E1ZB,GSTR1,2026-07,2026-27,2026-08-11,FILED,920000,165600,0,AA2907269876544
24AAACA4321C1Z2,GSTR3B,2026-07,2026-27,2026-08-20,PENDING,4500000,810000,560000,
24AAACA4321C1Z2,GSTR1,2026-07,2026-27,2026-08-11,PREPARED,4500000,810000,0,
27AABFM1122K1Z4,CMP08,2026-Q1,2026-27,2026-07-18,FILED,650000,6500,0,AA2707265544332
33AAACM6655B1Z1,GSTR3B,2026-07,2026-27,2026-08-20,PENDING,12500000,2250000,1800000,
"""

    with open("d:/Projects/Taxoryn/sample_gst_filings_migration.csv", "w", encoding="utf-8") as f:
        f.write(filings_csv)
    with open("d:/Projects/Taxoryn/frontend/public/sample_gst_filings_migration.csv", "w", encoding="utf-8") as f:
        f.write(filings_csv)

    # 3. Styled GST Profiles Excel Workbook (.xlsx)
    wb_prof = openpyxl.Workbook()
    ws_prof = wb_prof.active
    ws_prof.title = "Client GST Registrations"

    prof_headers = ["Client PAN", "GSTIN", "Legal Name", "Trade Name", "GST Scheme", "Filing Frequency", "Registration Date", "State Code"]
    ws_prof.append(prof_headers)

    prof_data = [
        ["AAACZ1234D", "27AAACZ1234D1Z8", "Zenith Infotech Solutions Private Limited", "Zenith Infotech", "REGULAR", "MONTHLY", "2018-04-01", "27"],
        ["AAALB5678E", "29AAALB5678E1ZB", "Bluecrest Logistics LLP", "Bluecrest Express", "REGULAR", "MONTHLY", "2019-07-15", "29"],
        ["ABCPJ9876M", "27ABCPJ9876M1ZX", "Anand Ramesh Joshi", "Joshi Tax Consultancy", "REGULAR", "QUARTERLY", "2021-11-01", "27"],
        ["AABFM1122K", "27AABFM1122K1Z4", "Mundeshwari Trading Co", "Mundeshwari Traders", "COMPOSITION", "QUARTERLY", "2017-07-01", "27"],
        ["AAACA4321C", "24AAACA4321C1Z2", "Apex Agro Food Products Limited", "Apex Foods", "REGULAR", "MONTHLY", "2018-09-20", "24"],
        ["AABFS7788P", "07AABFS7788P1Z6", "Shree Ganesh Enterprises", "Ganesh Retailers", "REGULAR", "MONTHLY", "2020-01-10", "07"],
        ["AAALS9988Q", "27AAALS9988Q1Z9", "Sunrise Healthcare LLP", "Sunrise Clinic & Diagnostics", "REGULAR", "MONTHLY", "2021-03-05", "27"],
        ["AAACM6655B", "33AAACM6655B1Z1", "Metro Infrastructure Developers Private Limited", "Metro Infra", "REGULAR", "MONTHLY", "2017-08-12", "33"],
    ]

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col_idx in range(1, len(prof_headers) + 1):
        cell = ws_prof.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws_prof.row_dimensions[1].height = 28

    for row_idx, row_values in enumerate(prof_data, start=2):
        ws_prof.append(row_values)
        ws_prof.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(prof_headers) + 1):
            cell = ws_prof.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if col_idx in [1, 2, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    for col in ws_prof.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_prof.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb_prof.save("d:/Projects/Taxoryn/Taxoryn_Sample_GST_Profiles_Migration.xlsx")
    wb_prof.save("d:/Projects/Taxoryn/frontend/public/Taxoryn_Sample_GST_Profiles_Migration.xlsx")

    # 4. Styled GST Filings Excel Workbook (.xlsx)
    wb_fil = openpyxl.Workbook()
    ws_fil = wb_fil.active
    ws_fil.title = "Historical GST Returns"

    fil_headers = ["GSTIN", "Return Type", "Return Period", "Financial Year", "Due Date", "Filing Status", "Taxable Value", "Tax Liability", "ITC Claimed", "ARN Number"]
    ws_fil.append(fil_headers)

    fil_data = [
        ["27AAACZ1234D1Z8", "GSTR3B", "2026-07", "2026-27", "2026-08-20", "FILED", 1850000, 333000, 240000, "AA2707261234567"],
        ["27AAACZ1234D1Z8", "GSTR1", "2026-07", "2026-27", "2026-08-11", "FILED", 1850000, 333000, 0, "AA2707261234568"],
        ["29AAALB5678E1ZB", "GSTR3B", "2026-07", "2026-27", "2026-08-20", "FILED", 920000, 165600, 110000, "AA2907269876543"],
        ["29AAALB5678E1ZB", "GSTR1", "2026-07", "2026-27", "2026-08-11", "FILED", 920000, 165600, 0, "AA2907269876544"],
        ["24AAACA4321C1Z2", "GSTR3B", "2026-07", "2026-27", "2026-08-20", "PENDING", 4500000, 810000, 560000, ""],
        ["24AAACA4321C1Z2", "GSTR1", "2026-07", "2026-27", "2026-08-11", "PREPARED", 4500000, 810000, 0, ""],
        ["27AABFM1122K1Z4", "CMP08", "2026-Q1", "2026-27", "2026-07-18", "FILED", 650000, 6500, 0, "AA2707265544332"],
        ["33AAACM6655B1Z1", "GSTR3B", "2026-07", "2026-27", "2026-08-20", "PENDING", 12500000, 2250000, 1800000, ""],
    ]

    for col_idx in range(1, len(fil_headers) + 1):
        cell = ws_fil.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws_fil.row_dimensions[1].height = 28

    for row_idx, row_values in enumerate(fil_data, start=2):
        ws_fil.append(row_values)
        ws_fil.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, len(fil_headers) + 1):
            cell = ws_fil.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if col_idx in [1, 2, 3, 4, 5, 6, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for col in ws_fil.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_fil.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb_fil.save("d:/Projects/Taxoryn/Taxoryn_Sample_GST_Filings_Migration.xlsx")
    wb_fil.save("d:/Projects/Taxoryn/frontend/public/Taxoryn_Sample_GST_Filings_Migration.xlsx")
    print("GST sample Excel and CSV created successfully!")

if __name__ == "__main__":
    create_sample_gst_files()
